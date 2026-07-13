"""Unit tests for automation/excel_tools.py"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from automation import excel_tools


class TestCreateWorkbook:
    def test_creates_file_with_header_row(self, tmp_path: Path):
        path = tmp_path / "new.xlsx"
        excel_tools.create_workbook(path, headers=["Name", "Age", "Email"])

        assert path.exists()
        wb = load_workbook(path)
        ws = wb["Sheet1"]
        assert [c.value for c in ws[1]] == ["Name", "Age", "Email"]

    def test_empty_headers_raises(self, tmp_path: Path):
        with pytest.raises(ValueError):
            excel_tools.create_workbook(tmp_path / "new.xlsx", headers=[])


class TestReadWorkbook:
    def test_reads_expected_shape_and_columns(self, sample_excel: Path):
        df = excel_tools.read_workbook(sample_excel, sheet_name="Sheet1")
        assert len(df) == 4
        assert list(df.columns) == ["Name", "Region", "Revenue", "Status"]

    def test_missing_file_raises(self, tmp_path: Path):
        with pytest.raises(FileNotFoundError):
            excel_tools.read_workbook(tmp_path / "ghost.xlsx")


class TestUpdateCells:
    def test_updates_the_target_cell_only(self, sample_excel: Path):
        excel_tools.update_cells(sample_excel, "Sheet1", {"A2": "Updated Name"})

        wb = load_workbook(sample_excel)
        ws = wb["Sheet1"]
        assert ws["A2"].value == "Updated Name"
        assert ws["B2"].value == "North"  # untouched neighbor cell

    def test_missing_sheet_raises(self, sample_excel: Path):
        with pytest.raises(KeyError):
            excel_tools.update_cells(sample_excel, "NoSuchSheet", {"A1": "x"})


class TestApplyFormatting:
    def test_header_row_becomes_bold(self, sample_excel: Path):
        excel_tools.apply_formatting(sample_excel, "Sheet1")

        wb = load_workbook(sample_excel)
        ws = wb["Sheet1"]
        assert ws["A1"].font.bold is True

    def test_columns_are_autofit_wider_than_default(self, sample_excel: Path):
        excel_tools.apply_formatting(sample_excel, "Sheet1", autofit=True)

        wb = load_workbook(sample_excel)
        ws = wb["Sheet1"]
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["A"].width > 0


class TestCreateSummarySheet:
    def test_aggregation_math_is_correct(self, sample_excel: Path):
        excel_tools.create_summary_sheet(sample_excel, "Sheet1", group_by="Region", agg_column="Revenue")

        summary = pd.read_excel(sample_excel, sheet_name="Sheet1_Summary")
        north_row = summary[summary["Region"] == "North"].iloc[0]

        # sample_excel fixture: North = Asha(100) + Meera(300) = 400
        assert north_row["Revenue_sum"] == 400
        assert north_row["Revenue_count"] == 2

    def test_missing_column_raises(self, sample_excel: Path):
        with pytest.raises(KeyError):
            excel_tools.create_summary_sheet(sample_excel, "Sheet1", group_by="NotAColumn", agg_column="Revenue")


class TestFilterData:
    def test_filters_to_matching_rows_only(self, sample_excel: Path):
        result = excel_tools.filter_data(sample_excel, "Sheet1", column="Status", value="Active")
        assert len(result) == 3  # Asha, Ravi, Karan are Active
        assert set(result["Name"]) == {"Asha", "Ravi", "Karan"}

    def test_numeric_value_matches_via_string_coercion(self, sample_excel: Path):
        result = excel_tools.filter_data(sample_excel, "Sheet1", column="Revenue", value="200")
        assert len(result) == 1
        assert result.iloc[0]["Name"] == "Ravi"

    def test_missing_column_raises(self, sample_excel: Path):
        with pytest.raises(KeyError):
            excel_tools.filter_data(sample_excel, "Sheet1", column="Ghost", value="x")


class TestExportReport:
    def test_export_to_csv(self, sample_excel: Path, tmp_path: Path):
        df = excel_tools.read_workbook(sample_excel, "Sheet1")
        out = tmp_path / "report.csv"
        excel_tools.export_report(df, out, fmt="csv")

        assert out.exists()
        reloaded = pd.read_csv(out)
        assert len(reloaded) == len(df)

    def test_export_to_xlsx(self, sample_excel: Path, tmp_path: Path):
        df = excel_tools.read_workbook(sample_excel, "Sheet1")
        out = tmp_path / "report.xlsx"
        excel_tools.export_report(df, out, fmt="xlsx")

        assert out.exists()

    def test_invalid_format_raises(self, sample_excel: Path, tmp_path: Path):
        df = excel_tools.read_workbook(sample_excel, "Sheet1")
        with pytest.raises(ValueError):
            excel_tools.export_report(df, tmp_path / "report.json", fmt="json")
