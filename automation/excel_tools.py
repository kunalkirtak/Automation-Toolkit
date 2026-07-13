"""
automation/excel_tools.py
============================
Excel automation built on top of ``openpyxl`` (for cell-level control:
creating sheets, updating cells, formatting) and ``pandas`` (for
data-level work: reading, filtering, grouping, exporting).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from automation.logger import get_logger
from config import default_config

logger = get_logger(__name__)


def _require_file(path: Path) -> Path:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return path


def read_workbook(path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Read a worksheet into a pandas DataFrame.

    If ``sheet_name`` is None, reads the first sheet.
    """
    path = _require_file(path)
    df = pd.read_excel(path, sheet_name=sheet_name or 0)
    logger.info("Read %d row(s) x %d column(s) from %s", len(df), len(df.columns), path.name)
    return df


def create_workbook(path: Path, headers: List[str], sheet_name: str = "Sheet1") -> Path:
    """Create a new .xlsx workbook with a single header row."""
    if not headers:
        raise ValueError("headers must contain at least one column name")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(path)

    logger.info("Created workbook %s with sheet '%s' (%d columns)", path, sheet_name, len(headers))
    return path


def update_cells(path: Path, sheet_name: str, updates: Dict[str, Any]) -> Path:
    """Update specific cells, e.g. {'B2': 'value'}, in an existing workbook."""
    path = _require_file(path)
    wb = load_workbook(path)

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in {path.name}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    for cell_ref, value in updates.items():
        ws[cell_ref] = value

    wb.save(path)
    logger.info("Updated %d cell(s) in %s!%s", len(updates), path.name, sheet_name)
    return path


def apply_formatting(
    path: Path, sheet_name: str, bold_header: bool = True, autofit: bool = True
) -> Path:
    """Apply header styling and column auto-sizing to a worksheet."""
    path = _require_file(path)
    wb = load_workbook(path)

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in {path.name}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    if bold_header and ws.max_row >= 1:
        fill_color = default_config.header_fill_color
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

    if autofit:
        for column_cells in ws.columns:
            lengths = [len(str(c.value)) for c in column_cells if c.value is not None]
            width = (max(lengths) if lengths else 8) + 2
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = width

    wb.save(path)
    logger.info("Applied formatting to %s!%s", path.name, sheet_name)
    return path


def create_summary_sheet(path: Path, source_sheet: str, group_by: str, agg_column: str) -> Path:
    """Append a summary sheet grouping ``source_sheet`` by ``group_by``.

    The summary reports sum, mean, and count of ``agg_column`` per
    group, written to a new sheet named '{source_sheet}_Summary'
    (replacing it if it already exists).
    """
    path = _require_file(path)
    df = pd.read_excel(path, sheet_name=source_sheet)

    for col in (group_by, agg_column):
        if col not in df.columns:
            raise KeyError(f"Column '{col}' not found in sheet '{source_sheet}'. Available: {list(df.columns)}")

    summary = (
        df.groupby(group_by)[agg_column]
        .agg(["sum", "mean", "count"])
        .reset_index()
        .rename(columns={"sum": f"{agg_column}_sum", "mean": f"{agg_column}_mean", "count": f"{agg_column}_count"})
    )

    summary_sheet_name = f"{source_sheet}_Summary"[:31]  # Excel's 31-char sheet name limit
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name=summary_sheet_name, index=False)

    logger.info(
        "Created summary sheet '%s' (%d groups) in %s", summary_sheet_name, len(summary), path.name
    )
    return path


def filter_data(path: Path, sheet_name: str, column: str, value: Any) -> pd.DataFrame:
    """Return rows from ``sheet_name`` where ``column`` equals ``value``.

    Comparison is done as strings so numeric CLI input (which always
    arrives as text) still matches numeric cell values.
    """
    path = _require_file(path)
    df = pd.read_excel(path, sheet_name=sheet_name)

    if column not in df.columns:
        raise KeyError(f"Column '{column}' not found in sheet '{sheet_name}'. Available: {list(df.columns)}")

    filtered = df[df[column].astype(str) == str(value)].reset_index(drop=True)
    logger.info("Filtered %s!%s: %d/%d row(s) match %s=%s", path.name, sheet_name, len(filtered), len(df), column, value)
    return filtered


def export_report(df: pd.DataFrame, output_path: Path, fmt: str = "xlsx") -> Path:
    """Export a DataFrame to ``output_path`` as 'xlsx' or 'csv'."""
    if fmt not in ("xlsx", "csv"):
        raise ValueError("fmt must be 'xlsx' or 'csv'")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if fmt == "csv":
        df.to_csv(output_path, index=False)
    else:
        df.to_excel(output_path, index=False)

    logger.info("Exported %d row(s) to %s", len(df), output_path)
    return output_path
